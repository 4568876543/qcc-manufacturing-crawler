# Excel操作工具函数
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def create_excel_template(output_file):
    """
    创建Excel模板文件，包含总览表结构
    
    Args:
        output_file (str): 输出的Excel文件路径
    """
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 创建总览表
    overview_sheet = wb.create_sheet(title="总览")
    overview_sheet['A1'] = "重庆市制造业企业统计"
    overview_sheet['A1'].font = Font(bold=True, size=14)
    overview_sheet.merge_cells('A1:C1')
    
    # 设置表头
    headers = ["区县", "行业类别", "企业数量"]
    for col, header in enumerate(headers, 1):
        cell = overview_sheet.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
    
    # 调整列宽
    overview_sheet.column_dimensions['A'].width = 20
    overview_sheet.column_dimensions['B'].width = 35
    overview_sheet.column_dimensions['C'].width = 12
    
    # 保存工作簿
    wb.save(output_file)
    print(f"[Excel] 模板已创建: {output_file}")


def update_excel_data(output_file, data):
    """
    更新Excel文件中的总览表数据
    
    Args:
        output_file (str): Excel文件路径
        data (list): 包含区县、行业类别和企业数量的数据列表
    """
    # 将数据转换为DataFrame
    new_data = pd.DataFrame(data, columns=["区县", "行业类别", "企业数量"])
    
    # 检查文件是否存在
    if os.path.exists(output_file):
        try:
            # 读取现有数据
            existing_data = pd.read_excel(output_file, sheet_name="总览", header=2)
            
            # 合并数据，保留最新的企业数量
            combined_data = pd.concat([existing_data, new_data], ignore_index=True)
            combined_data = combined_data.drop_duplicates(subset=["区县", "行业类别"], keep="last")
        except Exception as e:
            print(f"[Excel] 读取现有数据失败: {e}")
            combined_data = new_data
    else:
        combined_data = new_data
    
    # 按区县和行业类别排序
    combined_data = combined_data.sort_values(by=["区县", "行业类别"]).reset_index(drop=True)
    
    # 使用ExcelWriter写入
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_data.to_excel(writer, sheet_name="总览", index=False, startrow=2, header=True)
    
    print(f"[Excel] 总览表已更新: {len(combined_data)} 条记录")


def create_district_sheets(output_file, districts, industries):
    """
    为每个区县创建单独的工作表
    
    Args:
        output_file (str): Excel文件路径
        districts (list): 区县列表
        industries (dict): 行业代码和名称字典
    """
    if not os.path.exists(output_file):
        create_excel_template(output_file)
    
    # 读取现有Excel
    with pd.ExcelFile(output_file) as xl:
        existing_sheets = xl.sheet_names
    
    # 使用ExcelWriter追加工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # 复制现有工作表（跳过将创建的区县工作表）
        for sheet_name in existing_sheets:
            if sheet_name not in districts and sheet_name not in ["区县汇总", "行业汇总"]:
                df = pd.read_excel(output_file, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 为每个区县创建工作表
        for district in districts:
            # 创建区县数据表
            district_df = pd.DataFrame({
                "行业代码": list(industries.keys()),
                "行业名称": list(industries.values()),
                "企业数量": [0] * len(industries)
            })
            district_df.to_excel(writer, sheet_name=district, index=False)
    
    print(f"[Excel] 已创建 {len(districts)} 个区县工作表")


def update_district_sheet(output_file, district, industry_code, industry_name, count):
    """
    更新区县工作表中的数据
    
    Args:
        output_file (str): Excel文件路径
        district (str): 区县名称
        industry_code (str): 行业代码
        industry_name (str): 行业名称
        count (int): 企业数量
    """
    if not os.path.exists(output_file):
        print(f"[Excel] 文件不存在: {output_file}")
        return
    
    try:
        # 读取所有工作表
        all_sheets = {}
        with pd.ExcelFile(output_file) as xl:
            for sheet_name in xl.sheet_names:
                all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
        
        # 更新区县工作表
        if district in all_sheets:
            district_df = all_sheets[district]
            
            # 更新数据
            mask = district_df["行业代码"] == industry_code
            if mask.any():
                district_df.loc[mask, "企业数量"] = count
            else:
                # 添加新行
                new_row = pd.DataFrame({
                    "行业代码": [industry_code],
                    "行业名称": [industry_name],
                    "企业数量": [count]
                })
                district_df = pd.concat([district_df, new_row], ignore_index=True)
            
            all_sheets[district] = district_df
        
        # 写回所有工作表
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in all_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"[Excel] 更新 {district} - {industry_name}: {count} 家企业")
        
    except Exception as e:
        print(f"[Excel] 更新区县工作表失败: {e}")


def create_summary_sheet(output_file, districts, industries):
    """
    创建汇总工作表
    
    Args:
        output_file (str): Excel文件路径
        districts (list): 区县列表
        industries (dict): 行业代码和名称字典
    """
    if not os.path.exists(output_file):
        create_excel_template(output_file)
    
    # 创建按区县汇总的DataFrame
    district_summary = pd.DataFrame({
        "区县": districts,
        "企业总数": [0] * len(districts)
    })
    
    # 创建按行业汇总的DataFrame
    industry_summary = pd.DataFrame({
        "行业代码": list(industries.keys()),
        "行业名称": list(industries.values()),
        "企业总数": [0] * len(industries)
    })
    
    # 读取现有工作表
    all_sheets = {}
    if os.path.exists(output_file):
        with pd.ExcelFile(output_file) as xl:
            for sheet_name in xl.sheet_names:
                if sheet_name not in ["区县汇总", "行业汇总"]:
                    all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    
    # 写入所有工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        # 写入原有工作表
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 写入汇总工作表
        district_summary.to_excel(writer, sheet_name="区县汇总", index=False)
        industry_summary.to_excel(writer, sheet_name="行业汇总", index=False)
    
    print(f"[Excel] 汇总工作表已创建")


def update_summary_sheet(output_file, data):
    """
    更新汇总工作表
    
    Args:
        output_file (str): Excel文件路径
        data (list): 包含区县、行业类别和企业数量的数据列表
    """
    if not os.path.exists(output_file):
        print(f"[Excel] 文件不存在: {output_file}")
        return
    
    if not data:
        print(f"[Excel] 无数据可更新汇总表")
        return
    
    # 将数据转换为DataFrame
    df = pd.DataFrame(data)
    
    # 按区县汇总
    district_summary = df.groupby("区县")["企业数量"].sum().reset_index()
    district_summary.columns = ["区县", "企业总数"]
    
    # 按行业汇总
    industry_summary = df.groupby(["行业代码", "行业类别"])["企业数量"].sum().reset_index()
    industry_summary.columns = ["行业代码", "行业名称", "企业总数"]
    
    # 读取现有工作表
    all_sheets = {}
    with pd.ExcelFile(output_file) as xl:
        for sheet_name in xl.sheet_names:
            if sheet_name not in ["区县汇总", "行业汇总"]:
                all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    
    # 写入所有工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        # 写入原有工作表
        for sheet_name, df_sheet in all_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 写入更新后的汇总工作表
        district_summary.to_excel(writer, sheet_name="区县汇总", index=False)
        industry_summary.to_excel(writer, sheet_name="行业汇总", index=False)
    
    print(f"[Excel] 汇总表已更新: 区县 {len(district_summary)} 条, 行业 {len(industry_summary)} 条")


def update_all_district_sheets(output_file, data, districts, industries):
    """
    批量更新所有区县工作表
    
    Args:
        output_file (str): Excel文件路径
        data (list): 完整数据列表
        districts (list): 区县列表
        industries (dict): 行业字典
    """
    if not data:
        return
    
    df = pd.DataFrame(data)
    
    # 读取现有工作表
    all_sheets = {}
    with pd.ExcelFile(output_file) as xl:
        for sheet_name in xl.sheet_names:
            all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    
    # 更新每个区县工作表
    for district in districts:
        if district in all_sheets:
            district_df = all_sheets[district]
            
            # 获取该区县的数据
            district_data = df[df["区县"] == district]
            
            for _, row in district_data.iterrows():
                mask = district_df["行业代码"] == row["行业代码"]
                if mask.any():
                    district_df.loc[mask, "企业数量"] = row["企业数量"]
            
            all_sheets[district] = district_df
    
    # 写回所有工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        for sheet_name, df_sheet in all_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"[Excel] 已批量更新所有区县工作表")
